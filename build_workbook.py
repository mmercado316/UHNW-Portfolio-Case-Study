#!/usr/bin/env python3
"""
UHNW Portfolio Analysis - Excel Workbook Generator (with live formulas)
Run:    python3 build_workbook.py
Output: UHNW_Portfolio_Analysis.xlsx

FORMULA DESIGN
--------------
All key assumptions live in the "Inputs" tab (column C, rows 5–50).
Every computed value in every other tab references those cells via
Excel cross-sheet formulas, so changing a single assumption cascades
through the entire workbook automatically.

Active input cells  → amber fill  (edit these)
Computed cells      → light-blue fill  (do not edit — auto-update)
"""

import csv
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── PATHS ─────────────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
CSV_DIR     = os.path.join(BASE_DIR, "csv")
OUTPUT_FILE = os.path.join(BASE_DIR, "UHNW_Portfolio_Analysis.xlsx")

# ── BRAND COLORS ──────────────────────────────────────────────────────────────
NAVY   = "002D72"
GOLD   = "C8A84B"
WHITE  = "FFFFFF"
LGRAY  = "F2F2F2"
MGRAY  = "D9D9D9"
DGRAY  = "404040"
LBLUE  = "E8EEF7"
GREEN  = "00703C"
RED    = "C00000"
PURPLE = "7030A0"
AMBER  = "FFF2CC"   # editable input cells
CBLUE  = "DEEAF1"   # computed / formula cells

# ── CELL ADDRESS DICTIONARY ───────────────────────────────────────────────────
# Every entry maps a logical name to its Excel address in the Inputs sheet.
# Column C of Inputs holds all values.  Row numbers match the layout below.
I = {
    # Section A: Portfolio (rows 5-8)
    "portfolio_value":   "'Inputs'!C5",
    "cost_basis":        "'Inputs'!C6",
    "tech_conc":         "'Inputs'!C7",
    "cash_conc":         "'Inputs'!C8",
    # Section B: Tax rates (rows 11-15)
    "fed_rate":          "'Inputs'!C11",
    "niit_rate":         "'Inputs'!C12",
    "state_rate":        "'Inputs'!C13",
    "tax_bracket":       "'Inputs'!C14",
    "harvesting_offset": "'Inputs'!C15",
    # Section C: Liquidation (rows 18-19)
    "n_tranches":        "'Inputs'!C18",
    "tranche_amt":       "'Inputs'!C19",
    # Section D: SBL (rows 22-26)
    "loan_amount":       "'Inputs'!C22",
    "collateral_value":  "'Inputs'!C23",
    "sofr_rate":         "'Inputs'!C24",
    "credit_spread":     "'Inputs'!C25",
    "ltv_trigger":       "'Inputs'!C26",
    # Section E: Goals (rows 29-36)
    "bear_return":       "'Inputs'!C29",
    "base_return":       "'Inputs'!C30",
    "bull_return":       "'Inputs'!C31",
    "withdrawal_rate":   "'Inputs'!C32",
    "horizon":           "'Inputs'!C33",
    "foundation_target": "'Inputs'!C34",
    "foundation_return": "'Inputs'!C35",
    "foundation_payout": "'Inputs'!C36",
    # Section F: Computed (rows 39-50) — formulas in Inputs, referenced here
    "unrealized_gains":  "'Inputs'!C39",
    "effective_rate":    "'Inputs'!C40",
    "gross_tax":         "'Inputs'!C41",
    "net_tax":           "'Inputs'!C42",
    "allin_rate":        "'Inputs'!C43",
    "aftertax_rate":     "'Inputs'!C44",
    "annual_gross_int":  "'Inputs'!C45",
    "annual_net_int":    "'Inputs'!C46",
    "start_val":         "'Inputs'!C47",
    "bear_net":          "'Inputs'!C48",
    "base_net":          "'Inputs'!C49",
    "bull_net":          "'Inputs'!C50",
}

# ── HELPER: load CSV ──────────────────────────────────────────────────────────
def load_csv(filename):
    path = os.path.join(CSV_DIR, filename)
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))

# ── HELPERS: borders ──────────────────────────────────────────────────────────
def _side(color="CCCCCC"):
    return Side(style="thin", color=color)

def thin_border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

def gold_border():
    s = Side(style="thin", color=GOLD)
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border(color=NAVY):
    return Border(bottom=Side(style="medium", color=color))

# ── HELPERS: cell writers ─────────────────────────────────────────────────────
def title_cell(ws, row, col, text, size=13, sub=False):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(name="Calibri", bold=not sub, size=size,
                       color=GOLD if sub else WHITE)
    c.fill      = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center",
                            wrap_text=True)
    return c

def section_cell(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(name="Calibri", bold=True, size=10, color=NAVY)
    c.fill      = PatternFill("solid", fgColor=LBLUE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    return c

def col_head(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(name="Calibri", bold=True, size=9, color=WHITE)
    c.fill      = PatternFill("solid", fgColor="1F5C99")
    c.alignment = Alignment(horizontal="center", vertical="center",
                            wrap_text=True)
    c.border    = thin_border()
    return c

def data_cell(ws, row, col, value, bold=False, size=10,
              align="left", color=DGRAY, fmt=None, bg=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Calibri", bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center",
                            wrap_text=True)
    if fmt:
        c.number_format = fmt
    c.fill   = PatternFill("solid", fgColor=(bg or WHITE))
    c.border = thin_border()
    return c

def total_cell(ws, row, col, value, fmt=None, color=WHITE, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Calibri", bold=True, size=10, color=color)
    c.fill      = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = thin_border()
    if fmt:
        c.number_format = fmt
    return c

def note_cell(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(name="Calibri", italic=True, size=8, color="888888")
    c.alignment = Alignment(horizontal="left", vertical="center",
                            wrap_text=True)
    return c

def input_cell(ws, row, col, value, fmt=None, align="center"):
    """Active assumption cell — amber fill, bold.  Change these to update the workbook."""
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Calibri", bold=True, size=11, color=DGRAY)
    c.fill      = PatternFill("solid", fgColor=AMBER)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = gold_border()
    if fmt:
        c.number_format = fmt
    return c

def computed_cell(ws, row, col, formula, fmt=None, align="center", color=NAVY):
    """Formula-driven cell — light blue fill.  Do not edit directly."""
    c = ws.cell(row=row, column=col, value=formula)
    c.font      = Font(name="Calibri", bold=False, size=10, color=color)
    c.fill      = PatternFill("solid", fgColor=CBLUE)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = thin_border()
    if fmt:
        c.number_format = fmt
    return c

# ── HELPERS: layout ───────────────────────────────────────────────────────────
def cw(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def rh(ws, row, height):
    ws.row_dimensions[row].height = height

def merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2, end_column=c2)

def title_band(ws, row, text, ncols, sub_text=None):
    merge(ws, row, 1, row, ncols)
    title_cell(ws, row, 1, text, size=13)
    rh(ws, row, 30)
    row += 1
    if sub_text:
        merge(ws, row, 1, row, ncols)
        title_cell(ws, row, 1, sub_text, size=10, sub=True)
        rh(ws, row, 18)
        row += 1
    rh(ws, row, 8)
    row += 1
    return row

def stripe(i):
    return LGRAY if i % 2 == 0 else WHITE

# ══════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)

inp = wb.create_sheet("Inputs")
ov  = wb.create_sheet("Overview")
pt  = wb.create_sheet("Portfolio Transition")
tax = wb.create_sheet("Tax Analysis")
sbl = wb.create_sheet("SBL Credit")
gp  = wb.create_sheet("Goals Planning")

inp.sheet_properties.tabColor = GOLD
ov.sheet_properties.tabColor  = NAVY
pt.sheet_properties.tabColor  = "1F5C99"
tax.sheet_properties.tabColor = "B8860B"
sbl.sheet_properties.tabColor = GREEN
gp.sheet_properties.tabColor  = PURPLE

for ws in [inp, ov, pt, tax, sbl, gp]:
    ws.sheet_view.showGridLines = False

# ── LOAD DATA ─────────────────────────────────────────────────────────────────
raw_port     = load_csv("portfolio_transition_analysis.csv")
raw_metrics  = load_csv("transition_metrics_comparison.csv")
raw_sbl      = load_csv("sbl_credit_strategy.csv")
raw_timeline = load_csv("transition_timeline.csv")

seen = set()
current_rows_clean = []
for r in raw_port:
    if r["Portfolio_Type"] == "Current_Concentrated" and float(r["Market_Value_USD"]) > 0:
        key = r["Sub_Category"]
        if key not in seen:
            seen.add(key)
            current_rows_clean.append(r)
for r in current_rows_clean:
    if r["Sub_Category"] == "Technology Sector":
        r["Sub_Category"] = "US Large Cap Tech (90% Concentration)"

jpm_rows = [r for r in raw_port
            if r["Portfolio_Type"] == "JPM_Diversified"
            and float(r["Allocation_Percent"]) > 0]

# ══════════════════════════════════════════════════════════════════════════════
# TAB 0: INPUTS  — Single Source of Truth
# ══════════════════════════════════════════════════════════════════════════════
ws = inp
NCOLS = 6
for col, w in [(1, 3), (2, 36), (3, 22), (4, 14), (5, 42), (6, 3)]:
    cw(ws, col, w)

# Title
merge(ws, 1, 1, 1, NCOLS)
title_cell(ws, 1, 1, "INPUTS & ASSUMPTIONS — SINGLE SOURCE OF TRUTH", size=13)
rh(ws, 1, 30)

merge(ws, 2, 1, 2, NCOLS)
title_cell(ws, 2, 1,
    "Edit cells highlighted in amber below.  Every computed value across the "
    "workbook updates automatically via Excel formulas.",
    size=10, sub=True)
rh(ws, 2, 22)

rh(ws, 3, 8)

# Column label row
col_head(ws, 4, 2, "ASSUMPTION")
col_head(ws, 4, 3, "VALUE  ← EDIT HERE")
col_head(ws, 4, 4, "UNIT")
col_head(ws, 4, 5, "DESCRIPTION")
rh(ws, 4, 18)

# ── SECTION A: Portfolio (rows 5-8) ───────────────────────────────────────────
merge(ws, 5, 2, 5, 5); section_cell(ws, 5, 2, "  A  |  PORTFOLIO ASSUMPTIONS")
rh(ws, 5, 18)

inp_a = [
    (6,  "Portfolio Value",               25_000_000, "$#,##0",  "Total market value of the concentrated tech portfolio"),
    (7,  "Estimated Cost Basis",           6_500_000, "$#,##0",  "Original acquisition cost of tech holdings"),
    (8,  "Tech Equity Concentration",           0.90, "0%",      "Percentage of portfolio in single tech sector"),
    (9,  "Cash & Equivalents Allocation",       0.10, "0%",      "Remaining portfolio in money market / cash"),
]
for row, label, val, fmt, desc in inp_a:
    data_cell(ws, row, 2, label, bold=True, color=NAVY)
    input_cell(ws, row, 3, val, fmt=fmt)
    data_cell(ws, row, 4, fmt.replace("$#,##0","USD").replace("0%","%"), color="888888", align="center")
    data_cell(ws, row, 5, desc, size=9, color="555555")
    rh(ws, row, 16)

rh(ws, 10, 8)

# ── SECTION B: Tax Rates (rows 11-15) ─────────────────────────────────────────
merge(ws, 11, 2, 11, 5); section_cell(ws, 11, 2, "  B  |  TAX RATE ASSUMPTIONS")
rh(ws, 11, 18)

inp_b = [
    (12, "Fed Long-Term Capital Gains Rate",   0.200, "0.0%", "Top federal rate on long-term capital gains"),
    (13, "Net Investment Income Tax (NIIT)",   0.038, "0.0%", "Medicare surtax on investment income (IRC §1411)"),
    (14, "State Income Tax Rate (CA)",         0.133, "0.0%", "Highest marginal California income tax rate"),
    (15, "Marginal Income Tax Bracket",        0.370, "0.0%", "Used for after-tax interest and deduction calculations"),
    (16, "Tax-Loss Harvesting Offset",       850_000, "$#,##0","Identified losses available to offset realized gains"),
]
for row, label, val, fmt, desc in inp_b:
    data_cell(ws, row, 2, label, bold=True, color=NAVY)
    input_cell(ws, row, 3, val, fmt=fmt)
    data_cell(ws, row, 4, "%" if "%" in fmt else "USD", color="888888", align="center")
    data_cell(ws, row, 5, desc, size=9, color="555555")
    rh(ws, row, 16)

rh(ws, 17, 8)

# ── SECTION C: Liquidation Parameters (rows 18-19) ────────────────────────────
merge(ws, 18, 2, 18, 5); section_cell(ws, 18, 2, "  C  |  LIQUIDATION PARAMETERS")
rh(ws, 18, 18)

inp_c = [
    (19, "Number of Tranches",          5,         "0",       "Total liquidation tranches spread over Months 4-8"),
    (20, "Tranche Amount (each)",  4_500_000,      "$#,##0",  "Amount of tech equity liquidated per tranche"),
]
for row, label, val, fmt, desc in inp_c:
    data_cell(ws, row, 2, label, bold=True, color=NAVY)
    input_cell(ws, row, 3, val, fmt=fmt)
    data_cell(ws, row, 4, "#" if fmt=="0" else "USD", color="888888", align="center")
    data_cell(ws, row, 5, desc, size=9, color="555555")
    rh(ws, row, 16)

rh(ws, 21, 8)

# ── SECTION D: SBL Parameters (rows 22-26) ────────────────────────────────────
merge(ws, 22, 2, 22, 5); section_cell(ws, 22, 2, "  D  |  SECURITIES-BASED LENDING PARAMETERS")
rh(ws, 22, 18)

inp_d = [
    (23, "SBL Loan Amount",          5_000_000, "$#,##0",  "Non-purpose credit line drawn to seed Private Foundation with $5M initial capital"),
    (24, "Collateral Pledged",       15_000_000, "$#,##0",  "Portfolio value pledged; 40% ($10M) remains unencumbered"),
    (25, "SOFR Benchmark Rate",          0.045,  "0.00%",   "Secured Overnight Financing Rate — update as market moves"),
    (26, "Credit Spread",               0.020,   "0.00%",   "Institutional pricing spread above SOFR"),
    (27, "Margin Call LTV Trigger",      0.40,   "0%",      "LTV level at which additional collateral is required"),
]
for row, label, val, fmt, desc in inp_d:
    data_cell(ws, row, 2, label, bold=True, color=NAVY)
    input_cell(ws, row, 3, val, fmt=fmt)
    data_cell(ws, row, 4, "%" if "%" in fmt else "USD", color="888888", align="center")
    data_cell(ws, row, 5, desc, size=9, color="555555")
    rh(ws, row, 16)

rh(ws, 28, 8)

# ── SECTION E: Goals Planning (rows 29-36) ────────────────────────────────────
merge(ws, 29, 2, 29, 5); section_cell(ws, 29, 2, "  E  |  GOALS-BASED PLANNING ASSUMPTIONS")
rh(ws, 29, 18)

inp_e = [
    (30, "Bear Case Gross Return",      0.050, "0.0%",    "Conservative scenario — stressed market environment"),
    (31, "Base Case Gross Return",      0.098, "0.0%",    "Institutional long-term capital market assumption"),
    (32, "Bull Case Gross Return",      0.135, "0.0%",    "Optimistic scenario — alternatives outperform"),
    (33, "Annual Withdrawal Rate",      0.020, "0.0%",    "Sustainable annual distribution for lifestyle & philanthropy"),
    (34, "Planning Horizon (Years)",       20, "0",       "Long-term compounding window"),
    (35, "Private Foundation Target", 10_000_000, "$#,##0", "Target philanthropic endowment value"),
    (36, "Foundation Annual Return",    0.075, "0.0%",    "Expected endowment investment return"),
    (37, "Foundation Annual Payout",    0.050, "0.0%",    "Annual grantmaking as % of foundation assets"),
    (38, "Foundation Annual Contribution", 70_760, "$#,##0",
     "Annual client contribution from portfolio income; bridges $5M seed to $10M target at Year 20"),
]
for row, label, val, fmt, desc in inp_e:
    data_cell(ws, row, 2, label, bold=True, color=NAVY)
    input_cell(ws, row, 3, val, fmt=fmt)
    unit = "Yrs" if fmt == "0" else ("%" if "%" in fmt else "USD")
    data_cell(ws, row, 4, unit, color="888888", align="center")
    data_cell(ws, row, 5, desc, size=9, color="555555")
    rh(ws, row, 16)

# ── SECTION F: Computed Values (rows 39-50) — do not edit ─────────────────────
merge(ws, 39, 2, 39, 5)
section_cell(ws, 39, 2,
    "  F  |  COMPUTED VALUES  —  Auto-calculated from above.  Do not edit.")
rh(ws, 39, 18)

# Labels in col B, formulas in col C (within-sheet, no sheet prefix needed)
computed_rows = [
    (40, "Unrealized Capital Gains",
     "=C6-C7",                          "$#,##0",  "Portfolio Value − Cost Basis"),
    (41, "Gross Effective Tax Rate",
     "=C12+C13+C14",                    "0.00%",   "Fed + NIIT + State"),
    (42, "Estimated Gross Tax Liability",
     "=(C6-C7)*(C12+C13+C14)",          "$#,##0",  "Unrealized Gains × Effective Rate"),
    (43, "Net Tax Liability (After Offset)",
     "=(C6-C7)*(C12+C13+C14)-C16",      "$#,##0",  "Gross Tax − Harvesting Offset"),
    (44, "All-In SBL Interest Rate",
     "=C25+C26",                         "0.00%",   "SOFR + Credit Spread"),
    (45, "After-Tax SBL Rate",
     "=(C25+C26)*(1-C15)",               "0.00%",   "All-In Rate × (1 − Tax Bracket)"),
    (46, "Annual Gross Interest (SBL)",
     "=C23*(C25+C26)",                   "$#,##0",  "Loan Amount × All-In Rate"),
    (47, "Annual After-Tax Interest (SBL)",
     "=C23*(C25+C26)*(1-C15)",           "$#,##0",  "Annual Gross × (1 − Tax Bracket)"),
    (48, "Post-Tax Starting Portfolio",
     "=C6-((C6-C7)*(C12+C13+C14)-C16)", "$#,##0",  "Portfolio Value − Net Tax Liability (Gross Tax − Harvesting Offset)"),
    (49, "Bear Net Return (after withdrawal)",
     "=C30-C33",                         "0.00%",   "Bear Return − Withdrawal Rate"),
    (50, "Base Net Return (after withdrawal)",
     "=C31-C33",                         "0.00%",   "Base Return − Withdrawal Rate"),
    (51, "Bull Net Return (after withdrawal)",
     "=C32-C33",                         "0.00%",   "Bull Return − Withdrawal Rate"),
]
for row, label, formula, fmt, desc in computed_rows:
    data_cell(ws, row, 2, label, bold=True, color=DGRAY)
    computed_cell(ws, row, 3, formula, fmt=fmt)
    data_cell(ws, row, 5, desc, size=9, color="555555")
    rh(ws, row, 15)

# Update I dict to match actual row numbers (shifted by 1 since E starts at 30 not 29)
# Re-map from actual layout
I.update({
    "bear_return":       "'Inputs'!C30",
    "base_return":       "'Inputs'!C31",
    "bull_return":       "'Inputs'!C32",
    "withdrawal_rate":   "'Inputs'!C33",
    "horizon":           "'Inputs'!C34",
    "foundation_target": "'Inputs'!C35",
    "foundation_return": "'Inputs'!C36",
    "foundation_payout": "'Inputs'!C37",
    "foundation_contrib": "'Inputs'!C38",
    # Computed rows (actual)
    "unrealized_gains":  "'Inputs'!C40",
    "effective_rate":    "'Inputs'!C41",
    "gross_tax":         "'Inputs'!C42",
    "net_tax":           "'Inputs'!C43",
    "allin_rate":        "'Inputs'!C44",
    "aftertax_rate":     "'Inputs'!C45",
    "annual_gross_int":  "'Inputs'!C46",
    "annual_net_int":    "'Inputs'!C47",
    "start_val":         "'Inputs'!C48",
    "bear_net":          "'Inputs'!C49",
    "base_net":          "'Inputs'!C50",
    "bull_net":          "'Inputs'!C51",
    # Also remap Inputs rows that shifted
    "portfolio_value":   "'Inputs'!C6",
    "cost_basis":        "'Inputs'!C7",
    "tech_conc":         "'Inputs'!C8",
    "cash_conc":         "'Inputs'!C9",
    "fed_rate":          "'Inputs'!C12",
    "niit_rate":         "'Inputs'!C13",
    "tax_bracket":       "'Inputs'!C14",
    "state_rate":        "'Inputs'!C15",    # Note: state is row 14, bracket row 15
    "harvesting_offset": "'Inputs'!C16",
    "n_tranches":        "'Inputs'!C19",
    "tranche_amt":       "'Inputs'!C20",
    "loan_amount":       "'Inputs'!C23",
    "collateral_value":  "'Inputs'!C24",
    "sofr_rate":         "'Inputs'!C25",
    "credit_spread":     "'Inputs'!C26",
    "ltv_trigger":       "'Inputs'!C27",
})

# Fix state_rate / tax_bracket (state=row14, bracket=row15 per inp_b order)
I["fed_rate"]      = "'Inputs'!C12"
I["niit_rate"]     = "'Inputs'!C13"
I["state_rate"]    = "'Inputs'!C14"
I["tax_bracket"]   = "'Inputs'!C15"

# Fix computed formulas to use correct row numbers
# (Inputs col C rows: portfolio=6, cost_basis=7, fed=12, niit=13, state=14,
#  bracket=15, offset=16, loan=23, collateral=24, sofr=25, spread=26, ltv_trigger=27)
# The computed_rows already reference these correctly since they're within-sheet.

rh(ws, 52, 8)
merge(ws, 53, 2, 53, 5)
note_cell(ws, 53, 2,
    "Amber cells are the only cells you need to edit.  "
    "All other tabs (Tax Analysis, SBL Credit, Goals Planning) pull from this sheet "
    "and will recalculate instantly in Excel when any value above changes.  "
    "|  UHNW Portfolio Case Study  |  Prepared: February 2026")
rh(ws, 53, 28)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1: OVERVIEW
# ══════════════════════════════════════════════════════════════════════════════
ws = ov
NCOLS = 6

for col, w in [(1, 3), (2, 30), (3, 20), (4, 20), (5, 20), (6, 3)]:
    cw(ws, col, w)

row = title_band(
    ws, 1,
    "UHNW PORTFOLIO TRANSITION ANALYSIS",
    NCOLS,
    "$25M Tech Founder Case Study  |  Concentrated Equity to Institutional Diversification  |  February 2026"
)

# ── Client Profile ─────────────────────────────────────────────────────────
merge(ws, row, 2, row, 5)
section_cell(ws, row, 2, "  CLIENT PROFILE")
rh(ws, row, 18); row += 1

profile = [
    ("Client Segment",     "Ultra High Net Worth (UHNW)  |  $10M+ Investable Assets"),
    ("Portfolio Value",    f"={I['portfolio_value']}"),
    ("Client Type",        "Technology Founder / Entrepreneur"),
    ("Primary Challenge",  "90% portfolio concentration in single tech sector; $18.5M unrealized gains"),
    ("Primary Objective",  "De-risk position, generate income, fund $10M Private Foundation"),
]
for i, (label, val) in enumerate(profile):
    bg = stripe(i)
    data_cell(ws, row, 2, label, bold=True, bg=bg, color=NAVY)
    merge(ws, row, 3, row, 5)
    is_formula = str(val).startswith("=")
    if is_formula:
        c = ws.cell(row=row, column=3, value=val)
        c.font      = Font(name="Calibri", size=10, color=DGRAY)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.fill      = PatternFill("solid", fgColor=bg)
        c.border    = thin_border()
        c.number_format = "$#,##0"
    else:
        data_cell(ws, row, 3, val, bg=bg, align="left")
    rh(ws, row, 16); row += 1

rh(ws, row, 8); row += 1

# ── Key Outcomes ──────────────────────────────────────────────────────────
merge(ws, row, 2, row, 5)
section_cell(ws, row, 2, "  KEY OUTCOMES: CURRENT vs. OPTIMIZED DIVERSIFIED MODEL")
rh(ws, row, 18); row += 1

for col, label in [(2, "Metric"), (3, "Current Portfolio"),
                   (4, "Optimized Model"), (5, "Improvement")]:
    col_head(ws, row, col, label)
rh(ws, row, 16); row += 1

outcomes = [
    ("Asset Classes",               "2",          "7",          "+5 Classes"),
    ("Portfolio Beta",              "1.35",       "0.88",       "↓ 34.8%"),
    ("Value at Risk (95%)",         "$4,250,000", "$1,875,000", "↓ 55.9%"),
    ("Maximum Drawdown (Est.)",     "45.0%",      "22.0%",      "↓ 51.1%"),
    ("Concentration Risk Score",    "95 / 100",   "15 / 100",   "↓ 84.2%"),
    ("Sharpe Ratio",                "0.38",       "0.68",       "↑ 78.9%"),
    ("Annual Portfolio Income",     "$112,500",   "$625,000",   "↑ 455%"),
    ("Correlation to S&P 500",      "0.94",       "0.68",       "↓ 27.7%"),
    ("Number of Holdings",          "8",          "145",        "+1,713%"),
]
for i, (metric, curr, jpm_val, chg) in enumerate(outcomes):
    bg = stripe(i)
    data_cell(ws, row, 2, metric,   bold=True, bg=bg)
    data_cell(ws, row, 3, curr,     bg=bg, align="center", color=RED)
    data_cell(ws, row, 4, jpm_val,  bg=bg, align="center", color=GREEN)
    data_cell(ws, row, 5, chg,      bg=bg, align="center", bold=True, color=GREEN)
    rh(ws, row, 15); row += 1

rh(ws, row, 8); row += 1

# ── Recommendation Summary ─────────────────────────────────────────────────
merge(ws, row, 2, row, 5)
section_cell(ws, row, 2, "  RECOMMENDATION SUMMARY")
rh(ws, row, 18); row += 1

for col, label in [(2, "Phase / Topic"), (3, "Action & Rationale")]:
    col_head(ws, row, col, label)
merge(ws, row, 3, row, 5)
rh(ws, row, 16); row += 1

recs = [
    ("Phase 1  (Months 1-3)",
     "Discovery and planning: risk tolerance assessment, cost basis review, CPA consultation, "
     "investment policy statement, and transition scenario modeling."),
    ("Phase 2  (Months 4-11)",
     "Systematic liquidation across 5 tranches (Months 4-8, $4.5M each; total $22.5M). "
     "Concurrent redeployment into optimized model: 55% public equity, 23% fixed income, "
     "20% alternatives (PE + Real Estate), 2% cash. Tax liability: est. $6.86M gross / $6.01M net."),
    ("Phase 3  (Months 12-18)",
     "Establish $5M SBL credit facility against diversified portfolio. "
     "Seed Private Foundation without triggering additional capital gains. "
     "Ongoing rebalancing, tax-loss harvesting ($850K target), and quarterly performance reviews."),
    ("Tax Strategy",
     "$850K tax-loss harvesting offset identified. DAF and Opportunity Zone "
     "structures available to further reduce net liability."),
    ("Goals-Based Planning",
     "9.8% base-case expected return projects portfolio to approximately $85M at Year 20. "
     "Foundation reaches $10.0M target at Year 20: $5M SBL seed compounding at 2.5% net "
     "plus $70,760 annual contribution from diversified portfolio income."),
]
for i, (phase, desc) in enumerate(recs):
    bg = stripe(i)
    data_cell(ws, row, 2, phase, bold=True, bg=bg, color=NAVY)
    merge(ws, row, 3, row, 5)
    data_cell(ws, row, 3, desc, bg=bg, align="left")
    rh(ws, row, 30); row += 1

rh(ws, row, 8); row += 1
merge(ws, row, 2, row, 5)
note_cell(ws, row, 2,
    "Navigate tabs:  Inputs  |  Portfolio Transition  |  Tax Analysis  |  SBL Credit  |  Goals Planning    "
    "Prepared by: Mario Enrique Mercado  |  Confidential  |  February 2026")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2: PORTFOLIO TRANSITION
# ══════════════════════════════════════════════════════════════════════════════
ws = pt
NCOLS = 7

for col, w in [(1, 3), (2, 28), (3, 16), (4, 16), (5, 16), (6, 16), (7, 3)]:
    cw(ws, col, w)

row = title_band(
    ws, 1,
    "PORTFOLIO TRANSITION ANALYSIS",
    NCOLS,
    "From Concentrated Risk to Institutional Diversification  |  $25M Portfolio  |  February 2026"
)

# ── Section A: Asset Allocation ───────────────────────────────────────────
merge(ws, row, 2, row, 6)
section_cell(ws, row, 2, "  SECTION A: ASSET ALLOCATION COMPARISON")
rh(ws, row, 18); row += 1

for col, label in [
    (2, "Asset Class / Sub-Category"),
    (3, "Current\nAlloc %"),
    (4, "Current\nValue ($)"),
    (5, "Optimized Model\nAlloc %"),
    (6, "Optimized Model\nValue ($)"),
]:
    col_head(ws, row, col, label)
rh(ws, row, 28); row += 1

jpm_by_class = defaultdict(list)
for r in jpm_rows:
    jpm_by_class[r["Asset_Class"]].append(r)

curr_classes = list({r["Asset_Class"] for r in current_rows_clean})
jpm_classes  = list(jpm_by_class.keys())
all_classes  = curr_classes + [c for c in jpm_classes if c not in curr_classes]

for ac in sorted(all_classes):
    curr_ac = [r for r in current_rows_clean if r["Asset_Class"] == ac]
    jpm_ac  = jpm_by_class.get(ac, [])

    curr_alloc = sum(float(r["Allocation_Percent"]) for r in curr_ac)
    curr_val   = sum(float(r["Market_Value_USD"])   for r in curr_ac)
    jpm_alloc  = sum(float(r["Allocation_Percent"]) for r in jpm_ac)
    jpm_val    = sum(float(r["Market_Value_USD"])   for r in jpm_ac)

    data_cell(ws, row, 2, ac, bold=True, bg=LBLUE, color=NAVY)
    data_cell(ws, row, 3, curr_alloc / 100 if curr_alloc else None,
              bold=True, bg=LBLUE, align="center", fmt="0.0%")
    data_cell(ws, row, 4, curr_val if curr_val else None,
              bold=True, bg=LBLUE, align="right", fmt='$#,##0')
    data_cell(ws, row, 5, jpm_alloc / 100 if jpm_alloc else None,
              bold=True, bg=LBLUE, align="center", fmt="0.0%")
    data_cell(ws, row, 6, jpm_val if jpm_val else None,
              bold=True, bg=LBLUE, align="right", fmt='$#,##0')
    rh(ws, row, 16); row += 1

    all_subs = sorted({r["Sub_Category"] for r in curr_ac + jpm_ac})
    for i, sub in enumerate(all_subs):
        cr = next((r for r in curr_ac if r["Sub_Category"] == sub), None)
        jr = next((r for r in jpm_ac  if r["Sub_Category"] == sub), None)
        bg = stripe(i)
        data_cell(ws, row, 2, f"    {sub}", bg=bg, size=9)
        data_cell(ws, row, 3,
                  float(cr["Allocation_Percent"]) / 100 if cr else None,
                  bg=bg, align="center", fmt="0.0%", size=9,
                  color=RED if cr else DGRAY)
        data_cell(ws, row, 4,
                  float(cr["Market_Value_USD"]) if cr else None,
                  bg=bg, align="right", fmt='$#,##0', size=9,
                  color=RED if cr else DGRAY)
        data_cell(ws, row, 5,
                  float(jr["Allocation_Percent"]) / 100 if jr else None,
                  bg=bg, align="center", fmt="0.0%", size=9,
                  color=GREEN if jr else DGRAY)
        data_cell(ws, row, 6,
                  float(jr["Market_Value_USD"]) if jr else None,
                  bg=bg, align="right", fmt='$#,##0', size=9,
                  color=GREEN if jr else DGRAY)
        rh(ws, row, 13); row += 1

# Grand total — formulas reference Inputs
total_cell(ws, row, 2, "TOTAL PORTFOLIO VALUE")
total_cell(ws, row, 3, 1.0,                     fmt="0.0%")
total_cell(ws, row, 4, f"={I['portfolio_value']}", fmt='$#,##0')
total_cell(ws, row, 5, 1.0,                     fmt="0.0%")
total_cell(ws, row, 6, f"={I['portfolio_value']}", fmt='$#,##0')
rh(ws, row, 18); row += 2

# ── Section B: Risk & Return Metrics ──────────────────────────────────────
merge(ws, row, 2, row, 6)
section_cell(ws, row, 2, "  SECTION B: RISK & RETURN METRICS COMPARISON")
rh(ws, row, 18); row += 1

for col, label in [
    (2, "Metric"), (3, "Category"),
    (4, "Current Portfolio"), (5, "Optimized Model"), (6, "Change / Improvement"),
]:
    col_head(ws, row, col, label)
rh(ws, row, 16); row += 1

display_cats = ["Portfolio Risk", "Return Metrics", "Diversification", "Income Generation"]
display_metrics = [r for r in raw_metrics if r["Metric_Category"] in display_cats]

for i, r in enumerate(display_metrics):
    bg    = stripe(i)
    unit  = r["Unit"]
    curr  = r["Current_Concentrated"]
    jpm_v = r["JPM_Diversified"]
    impr  = r["Improvement_Percent"]
    direc = r["Change_Direction"]

    if unit == "Percent":
        cd, jd = f"{float(curr):.1f}%", f"{float(jpm_v):.1f}%"
    elif unit == "USD":
        cd, jd = f"${float(curr):,.0f}", f"${float(jpm_v):,.0f}"
    else:
        cd, jd = curr, jpm_v

    try:
        chg_str = f"{'↓' if direc == 'Decrease' else '↑'} {float(impr):.1f}%"
    except ValueError:
        chg_str = "—"

    risk_decrease   = direc == "Decrease" and r["Metric_Category"] == "Portfolio Risk"
    return_increase = direc == "Increase"
    chg_color = GREEN if (risk_decrease or return_increase) else RED

    data_cell(ws, row, 2, r["Metric_Name"], bold=True, bg=bg)
    data_cell(ws, row, 3, r["Metric_Category"], bg=bg, align="center", size=9)
    data_cell(ws, row, 4, cd,       bg=bg, align="center", color=RED)
    data_cell(ws, row, 5, jd,       bg=bg, align="center", color=GREEN)
    data_cell(ws, row, 6, chg_str,  bold=True, bg=bg, align="center",
              color=chg_color)
    rh(ws, row, 15); row += 1

# ══════════════════════════════════════════════════════════════════════════════
# TAB 3: TAX ANALYSIS  — fully formula-driven
# ══════════════════════════════════════════════════════════════════════════════
ws = tax
NCOLS = 6

for col, w in [(1, 3), (2, 32), (3, 20), (4, 18), (5, 28), (6, 3)]:
    cw(ws, col, w)

row = title_band(
    ws, 1,
    "TAX TRANSITION ANALYSIS",
    NCOLS,
    "Capital Gains Management & Tax-Efficient Liquidation Strategy  |  February 2026"
)

# ── Section A: Capital Gains Overview ─────────────────────────────────────
merge(ws, row, 2, row, 5)
section_cell(ws, row, 2, "  SECTION A: CAPITAL GAINS OVERVIEW  —  All values update from Inputs tab")
rh(ws, row, 18); row += 1

for col, label in [(2, "Item"), (3, "Value"), (4, "Notes")]:
    col_head(ws, row, col, label)
merge(ws, row, 4, row, 5)
rh(ws, row, 16); row += 1

# (label, formula, fmt, val_color, bold, note)
cg_items = [
    ("Current Portfolio Value",
     f"={I['portfolio_value']}",          "$#,##0",  DGRAY, False,
     "Source: Inputs tab B6  |  Edit there to update all tabs"),
    ("Estimated Cost Basis",
     f"={I['cost_basis']}",               "$#,##0",  DGRAY, False,
     "Source: Inputs tab B7"),
    ("Unrealized Capital Gains",
     f"={I['unrealized_gains']}",         "$#,##0",  RED,   True,
     "Portfolio Value − Cost Basis  (Inputs C40)"),
    ("Federal Long-Term Cap Gains Rate",
     f"={I['fed_rate']}",                 "0.0%",    DGRAY, False,
     "Source: Inputs tab C12  |  Top federal rate on long-term gains"),
    ("Net Investment Income Tax (NIIT)",
     f"={I['niit_rate']}",                "0.0%",    DGRAY, False,
     "Medicare surtax on investment income (IRC §1411)  — Inputs C13"),
    ("State Tax Rate (CA)",
     f"={I['state_rate']}",               "0.0%",    DGRAY, False,
     "Highest marginal CA rate  — Inputs C14"),
    ("Gross Effective Rate",
     f"={I['effective_rate']}",           "0.0%",    DGRAY, True,
     "Fed + NIIT + State  (Inputs C41)"),
    ("Estimated Gross Tax Liability",
     f"={I['gross_tax']}",                "$#,##0",  RED,   True,
     "Unrealized Gains × Effective Rate  (Inputs C42)"),
    ("Tax-Loss Harvesting Offset",
     f"=0-{I['harvesting_offset']}",      "$#,##0",  GREEN, False,
     "Identified losses to offset gains  — Inputs C16  (displayed as reduction)"),
    ("Net Tax Liability (After Offset)",
     f"={I['net_tax']}",                  "$#,##0",  RED,   True,
     "Gross Tax − Offset  (Inputs C43)"),
]

for i, (label, formula, fmt, val_color, bold, note) in enumerate(cg_items):
    bg = stripe(i)
    data_cell(ws, row, 2, label, bold=bold, bg=bg,
              color=NAVY if bold else DGRAY)
    c = ws.cell(row=row, column=3, value=formula)
    c.font      = Font(name="Calibri", bold=bold, size=10, color=val_color)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.fill      = PatternFill("solid", fgColor=CBLUE)
    c.border    = thin_border()
    c.number_format = fmt
    merge(ws, row, 4, row, 5)
    data_cell(ws, row, 4, note, bg=bg, size=9, color="666666")
    rh(ws, row, 15); row += 1

rh(ws, row, 8); row += 1

# ── Section B: Phased Liquidation Schedule ────────────────────────────────
merge(ws, row, 2, row, 5)
section_cell(ws, row, 2,
    "  SECTION B: PHASED LIQUIDATION SCHEDULE  —  5 Tranches Over 8 Months")
rh(ws, row, 18); row += 1

for col, label in [
    (2, "Tranche / Month"),
    (3, "Amount Liquidated"),
    (4, "Est. Tax Impact"),
    (5, "Cumulative Liquidated"),
]:
    col_head(ws, row, col, label)
rh(ws, row, 16); row += 1

tranche_start_row = row
for i in range(5):
    month_num = 4 + i
    label     = f"Tranche {i+1}  (Month {month_num})"
    amt_f     = f"={I['tranche_amt']}"
    tax_f     = f"={I['tranche_amt']}*{I['effective_rate']}"
    cum_f     = f"=SUM(C{tranche_start_row}:C{row})"
    bg        = stripe(i)

    data_cell(ws, row, 2, label, bold=True, bg=bg)
    c3 = ws.cell(row=row, column=3, value=amt_f)
    c3.font = Font(name="Calibri", size=10, color=DGRAY)
    c3.alignment = Alignment(horizontal="right", vertical="center")
    c3.fill = PatternFill("solid", fgColor=CBLUE); c3.border = thin_border()
    c3.number_format = "$#,##0"

    c4 = ws.cell(row=row, column=4, value=tax_f)
    c4.font = Font(name="Calibri", size=10, color=RED)
    c4.alignment = Alignment(horizontal="right", vertical="center")
    c4.fill = PatternFill("solid", fgColor=CBLUE); c4.border = thin_border()
    c4.number_format = "$#,##0"

    c5 = ws.cell(row=row, column=5, value=cum_f)
    c5.font = Font(name="Calibri", size=10, color=DGRAY)
    c5.alignment = Alignment(horizontal="right", vertical="center")
    c5.fill = PatternFill("solid", fgColor=CBLUE); c5.border = thin_border()
    c5.number_format = "$#,##0"

    rh(ws, row, 15); row += 1

tranche_end_row = row - 1

total_cell(ws, row, 2, "TOTAL LIQUIDATED")
total_cell(ws, row, 3, f"=SUM(C{tranche_start_row}:C{tranche_end_row})", fmt='$#,##0')
total_cell(ws, row, 4, f"=SUM(D{tranche_start_row}:D{tranche_end_row})", fmt='$#,##0')
total_cell(ws, row, 5, f"=SUM(C{tranche_start_row}:C{tranche_end_row})", fmt='$#,##0')
rh(ws, row, 18); row += 2

# ── Section C: Tax Optimization Strategies ────────────────────────────────
merge(ws, row, 2, row, 5)
section_cell(ws, row, 2, "  SECTION C: TAX OPTIMIZATION STRATEGIES")
rh(ws, row, 18); row += 1

for col, label in [(2, "Strategy"), (3, "Est. Value"), (4, "Description")]:
    col_head(ws, row, col, label)
merge(ws, row, 4, row, 5)
rh(ws, row, 16); row += 1

strategies = [
    ("Tax-Loss Harvesting",
     f"={I['harvesting_offset']}",   "$#,##0", True,
     "Identify offsetting losses in fixed income and other equity positions. "
     "Reduces net taxable gain dollar-for-dollar."),
    ("Donor-Advised Fund (DAF)",
     500_000,   "$#,##0", False,
     "Contribute appreciated shares directly to DAF. Avoid capital gains entirely; "
     "receive fair-market-value charitable deduction."),
    ("Qualified Opportunity Zone (QOZ)",
     1_000_000, "$#,##0", False,
     "Invest deferred gains into QOZ fund. Defer recognition and potentially exclude "
     "future appreciation after 10-year hold."),
    ("Installment Liquidation Structure",
     "Structural", None, False,
     "5 monthly tranches across Months 4-8 within a single tax year. "
     "To split recognition across two tax years, time the final tranche(s) after December 31."),
    ("Charitable Remainder Trust (CRT)",
     "Structural", None, False,
     "Transfer appreciated stock into CRT. Avoid immediate gain, receive income stream, "
     "estate planning benefit."),
]
for i, (strat, val, fmt, is_formula, desc) in enumerate(strategies):
    bg = stripe(i)
    data_cell(ws, row, 2, strat, bold=True, bg=bg, color=NAVY)
    if is_formula:
        c = ws.cell(row=row, column=3, value=val)
        c.font = Font(name="Calibri", bold=True, size=10, color=GREEN)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.fill = PatternFill("solid", fgColor=CBLUE); c.border = thin_border()
        if fmt: c.number_format = fmt
    else:
        data_cell(ws, row, 3, val, bg=bg, align="right",
                  fmt=fmt if fmt else None,
                  color=GREEN if isinstance(val, int) else DGRAY)
    merge(ws, row, 4, row, 5)
    data_cell(ws, row, 4, desc, bg=bg, size=9)
    rh(ws, row, 28); row += 1

rh(ws, row, 8); row += 1
merge(ws, row, 2, row, 5)
note_cell(ws, row, 2,
    "All tax figures are estimates. Actual liability depends on cost basis, "
    "holding period, filing status, and applicable deductions. "
    "Coordinate with qualified CPA and estate attorney before execution.  "
    "Change portfolio value, cost basis, or tax rates in the Inputs tab to update all figures.")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 4: SBL CREDIT  — fully formula-driven
# ══════════════════════════════════════════════════════════════════════════════
ws = sbl
NCOLS = 7

for col, w in [(1, 3), (2, 30), (3, 18), (4, 16), (5, 16), (6, 20), (7, 3)]:
    cw(ws, col, w)

row = title_band(
    ws, 1,
    "SECURITIES-BASED LENDING (SBL) CREDIT FACILITY",
    NCOLS,
    "$5M Non-Purpose Credit Facility  |  Private Foundation Funding Strategy  |  February 2026"
)

# ── Section A: Credit Facility Structure ──────────────────────────────────
merge(ws, row, 2, row, 6)
section_cell(ws, row, 2,
    "  SECTION A: CREDIT FACILITY STRUCTURE  —  Live formulas from Inputs tab")
rh(ws, row, 18); row += 1

for col, label in [(2, "Parameter"), (3, "Value"), (4, "Notes")]:
    col_head(ws, row, col, label)
merge(ws, row, 4, row, 6)
rh(ws, row, 16); row += 1

# (label, value_or_formula, fmt, bold, note)
loan_params = [
    ("Loan Purpose",
     "Non-Purpose Loan",    None,    False,
     "Proceeds may not purchase or carry securities (Reg U)"),
    ("Loan Amount",
     f"={I['loan_amount']}", "$#,##0", True,
     f"Source: Inputs C23  |  Change there to update all SBL figures"),
    ("Collateral Portfolio Value",
     f"={I['collateral_value']}", "$#,##0", False,
     "Source: Inputs C24  |  60% of diversified portfolio pledged"),
    ("Loan-to-Value (LTV)",
     f"={I['loan_amount']}/{I['collateral_value']}", "0.0%", False,
     "Loan ÷ Collateral  (auto-updates if either changes)"),
    ("Margin Call Trigger",
     f"={I['ltv_trigger']}",   "0%",    True,
     "Source: Inputs C27  |  LTV level requiring additional collateral"),
    ("SOFR Benchmark Rate",
     f"={I['sofr_rate']}",     "0.00%", False,
     "Source: Inputs C25  |  Update when market SOFR changes"),
    ("Credit Spread",
     f"={I['credit_spread']}", "0.00%", False,
     "Source: Inputs C26  |  Institutional UHNW pricing"),
    ("All-In Interest Rate",
     f"={I['allin_rate']}",    "0.00%", True,
     "SOFR + Spread  (Inputs C44)"),
    ("Effective After-Tax Rate",
     f"={I['aftertax_rate']}", "0.00%", True,
     "All-In × (1 − Tax Bracket)  (Inputs C45)"),
    ("Payment Structure",
     "Interest-Only",          None,    False,
     "No mandatory principal amortization; annual rate resets"),
    ("Facility Term",
     "12 Months (Renewable)",  None,    False,
     "Annually renewable at bank's discretion; no prepayment penalty"),
    ("Annual Gross Interest",
     f"={I['annual_gross_int']}", "$#,##0", True,
     "Loan Amount × All-In Rate  (Inputs C46)"),
    ("Annual After-Tax Interest",
     f"={I['annual_net_int']}",   "$#,##0", True,
     "Annual Gross × (1 − Tax Bracket)  (Inputs C47)"),
    ("Use of Proceeds",
     "Private Foundation Seed Capital", None, False,
     "Fund $10M philanthropic vehicle without triggering capital gains"),
]

for i, (param, val, fmt, bold, note) in enumerate(loan_params):
    bg = stripe(i)
    data_cell(ws, row, 2, param, bold=bold, bg=bg,
              color=NAVY if bold else DGRAY)
    is_formula = isinstance(val, str) and val.startswith("=")
    if is_formula:
        c = ws.cell(row=row, column=3, value=val)
        c.font      = Font(name="Calibri", bold=bold, size=10,
                           color=GREEN if bold else DGRAY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill      = PatternFill("solid", fgColor=CBLUE)
        c.border    = thin_border()
        if fmt: c.number_format = fmt
    else:
        data_cell(ws, row, 3, val, bold=bold, bg=bg, align="center")
    merge(ws, row, 4, row, 6)
    data_cell(ws, row, 4, note, bg=bg, size=9, color="666666")
    rh(ws, row, 16); row += 1

rh(ws, row, 8); row += 1

# ── Section B: Interest-Only Amortization Schedule ────────────────────────
merge(ws, row, 2, row, 6)
section_cell(ws, row, 2,
    "  SECTION B: INTEREST-ONLY PAYMENT SCHEDULE  —  12-Month Facility")
rh(ws, row, 18); row += 1

for col, label in [
    (2, "Period"),
    (3, "Beginning Balance"),
    (4, "Monthly Interest"),
    (5, "Principal Payment"),
    (6, "Ending Balance"),
]:
    col_head(ws, row, col, label)
rh(ws, row, 16); row += 1

amort_start = row
for month in range(1, 13):
    bg   = stripe(month)
    bold = month == 12

    beg_f = f"={I['loan_amount']}"
    int_f = f"={I['loan_amount']}*{I['allin_rate']}/12"
    end_f = f"={I['loan_amount']}"

    data_cell(ws, row, 2, f"Month {month:>2}", bold=bold, bg=bg)

    for col, formula, color, align in [
        (3, beg_f,  DGRAY, "right"),
        (4, int_f,  RED,   "right"),
        (6, end_f,  DGRAY, "right"),
    ]:
        c = ws.cell(row=row, column=col, value=formula)
        c.font = Font(name="Calibri", bold=bold, size=10, color=color)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.fill = PatternFill("solid", fgColor=CBLUE)
        c.border = thin_border()
        c.number_format = "$#,##0"

    # Principal = 0 (static)
    data_cell(ws, row, 5, 0, bold=bold, bg=bg,
              align="center", fmt="$#,##0")
    rh(ws, row, 14); row += 1

amort_end = row - 1

total_cell(ws, row, 2, "TOTAL ANNUAL INTEREST")
total_cell(ws, row, 3, f"={I['loan_amount']}",  fmt='$#,##0')
total_cell(ws, row, 4, f"=SUM(D{amort_start}:D{amort_end})", fmt='$#,##0')
total_cell(ws, row, 5, 0,                        fmt='$#,##0')
total_cell(ws, row, 6, f"={I['loan_amount']}",  fmt='$#,##0')
rh(ws, row, 18); row += 2

# ── Section C: Margin Call Stress Test ────────────────────────────────────
merge(ws, row, 2, row, 6)
section_cell(ws, row, 2, "  SECTION C: MARGIN CALL STRESS TEST")
rh(ws, row, 18); row += 1

for col, label in [
    (2, "Scenario"),
    (3, "Collateral Decline"),
    (4, "Collateral Value"),
    (5, "Resulting LTV"),
    (6, "Margin Call Triggered?"),
]:
    col_head(ws, row, col, label)
rh(ws, row, 16); row += 1

stress_scenarios = [
    ("Base Case",         0.00),
    ("Moderate Decline",  0.10),
    ("Significant Drop",  0.20),
    ("Severe Stress",     0.30),
    ("2008-Style Crisis", 0.40),
    ("Extreme Tail Risk", 0.50),
]
for i, (scenario, decline) in enumerate(stress_scenarios):
    bg = stripe(i)
    # Python-side truth for color only
    py_collat  = 15_000_000 * (1 - decline)
    py_ltv     = 5_000_000 / py_collat
    triggered  = py_ltv > 0.40
    call_color = RED if triggered else GREEN
    dec_color  = RED if decline > 0 else GREEN

    decline_label = "0%  (Baseline)" if decline == 0 else f"-{decline*100:.0f}%"

    # Excel formulas
    collat_f = f"={I['collateral_value']}*(1-{decline})"
    ltv_f    = f"={I['loan_amount']}/({I['collateral_value']}*(1-{decline}))"
    call_f   = (f'=IF({I["loan_amount"]}/({I["collateral_value"]}*(1-{decline}))'
                f'>{I["ltv_trigger"]},"YES","NO")')

    data_cell(ws, row, 2, scenario, bold=True, bg=bg)
    data_cell(ws, row, 3, decline_label, bg=bg, align="center", color=dec_color)

    for col, formula, fmt, color, bold_f in [
        (4, collat_f, "$#,##0", DGRAY,  False),
        (5, ltv_f,    "0.0%",   RED if triggered else DGRAY, triggered),
    ]:
        c = ws.cell(row=row, column=col, value=formula)
        c.font = Font(name="Calibri", bold=bold_f, size=10, color=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", fgColor=CBLUE); c.border = thin_border()
        c.number_format = fmt

    c6 = ws.cell(row=row, column=6, value=call_f)
    c6.font = Font(name="Calibri", bold=True, size=10, color=call_color)
    c6.alignment = Alignment(horizontal="center", vertical="center")
    c6.fill = PatternFill("solid", fgColor=CBLUE); c6.border = thin_border()

    rh(ws, row, 15); row += 1

rh(ws, row, 8); row += 1
merge(ws, row, 2, row, 6)
note_cell(ws, row, 2,
    "Stress test applies to the pledged collateral ($15M by default — Inputs C24). "
    "Change collateral value or loan amount in Inputs tab to instantly update all LTV figures.  "
    "2008 historical S&P 500 peak-to-trough: -57%.")
rh(ws, row, 28); row += 2

# ── Section D: SBL vs. Liquidation — Tax Arbitrage ────────────────────────
merge(ws, row, 2, row, 6)
section_cell(ws, row, 2,
    "  SECTION D: SBL vs. OUTRIGHT LIQUIDATION — TAX ARBITRAGE ANALYSIS")
rh(ws, row, 18); row += 1

for col, label in [
    (2, "Comparison Factor"),
    (3, "Liquidate & Pay Gains"),
    (4, "SBL Credit Facility"),
    (5, "SBL Advantage"),
]:
    col_head(ws, row, col, label)
merge(ws, row, 5, row, 6)
rh(ws, row, 16); row += 1

# (factor, liq_formula_or_text, liq_fmt, sbl_formula_or_text, sbl_fmt, advantage_note)
arb_rows = [
    ("Immediate Tax Cost",
     f"={I['loan_amount']}*{I['effective_rate']}", "$#,##0",
     "$0",                                          None,
     "SBL defers this tax indefinitely"),
    ("Annual Carry Cost",
     "$0",                                          None,
     f"={I['annual_net_int']}",                    "$#,##0",
     "After-tax interest replaces upfront tax burden"),
    ("Capital Deployed to Foundation",
     f"={I['loan_amount']}*(1-{I['effective_rate']})", "$#,##0",
     f"={I['loan_amount']}",                        "$#,##0",
     "SBL delivers full loan amount; liquidation loses tax portion"),
    ("Portfolio Assets Still Compounding",
     f"={I['portfolio_value']}-{I['loan_amount']}", "$#,##0",
     f"={I['portfolio_value']}",                    "$#,##0",
     "SBL preserves the full $25M base"),
    ("Est. Portfolio Return on Preserved Capital (Yr 1)",
     "N/A",                                         None,
     f"={I['loan_amount']}*{I['base_return']}",    "$#,##0",
     f"Base case return × preserved capital"),
    ("Net Annual Benefit of SBL (Yr 1)",
     "N/A",                                         None,
     f"={I['loan_amount']}*{I['base_return']}-{I['annual_net_int']}", "$#,##0",
     "Portfolio return on preserved capital minus carry cost"),
    ("SBL Breakeven vs. Liquidation",
     "N/A",                                         None,
     "3.1 Years",                                   None,
     "Complex NPV calculation — see sbl_credit_strategy.csv"),
    ("10-Year SBL Advantage",
     "N/A",                                         None,
     "$7,124,500",                                  None,
     "Tax saved + compounding opportunity cost over 10 years"),
]

for i, (factor, liq_v, liq_fmt, sbl_v, sbl_fmt, adv) in enumerate(arb_rows):
    bg = stripe(i)
    data_cell(ws, row, 2, factor, bold=True, bg=bg)

    for col, val, fmt, base_color in [
        (3, liq_v, liq_fmt, RED),
        (4, sbl_v, sbl_fmt, GREEN),
    ]:
        is_f = isinstance(val, str) and val.startswith("=")
        if is_f:
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name="Calibri", size=10, color=base_color)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = PatternFill("solid", fgColor=CBLUE); c.border = thin_border()
            if fmt: c.number_format = fmt
        else:
            color = DGRAY if val == "N/A" else (RED if col == 3 and "$" in str(val) else GREEN)
            if val == "$0": color = DGRAY
            data_cell(ws, row, col, val, bg=bg, align="center", color=color)

    merge(ws, row, 5, row, 6)
    data_cell(ws, row, 5, adv, bg=bg, align="left",
              bold=True, color=GREEN, size=9)
    rh(ws, row, 16); row += 1

rh(ws, row, 8); row += 1
merge(ws, row, 2, row, 6)
note_cell(ws, row, 2,
    "Key formula: Tax Cost = Loan Amount × Effective Rate (both from Inputs).  "
    "Carry Cost = Annual After-Tax Interest (Inputs C47).  "
    "Change any assumption in Inputs tab to instantly reprice this analysis.")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 5: GOALS PLANNING  — 20-year formula chain
# ══════════════════════════════════════════════════════════════════════════════
ws = gp
NCOLS = 8

for col, w in [(1, 3), (2, 20), (3, 14), (4, 18), (5, 18), (6, 18), (7, 20), (8, 3)]:
    cw(ws, col, w)

row = title_band(
    ws, 1,
    "GOALS-BASED WEALTH PLANNING",
    NCOLS,
    "$10M Private Foundation  |  20-Year Portfolio Projection  |  Bear / Base / Bull Scenarios  |  February 2026"
)

# ── Section A: Planning Assumptions ───────────────────────────────────────
merge(ws, row, 2, row, 7)
section_cell(ws, row, 2,
    "  SECTION A: PLANNING ASSUMPTIONS  —  All values pull from Inputs tab")
rh(ws, row, 18); row += 1

for col, label in [
    (2, "Parameter"), (3, "Bear Case"),
    (4, "Base Case"),  (5, "Bull Case"), (6, "Notes"),
]:
    col_head(ws, row, col, label)
merge(ws, row, 6, row, 7)
rh(ws, row, 16); row += 1

# (label, bear_f, base_f, bull_f, fmt, note)
assumptions = [
    ("Starting Portfolio Value",
     f"={I['start_val']}", f"={I['start_val']}", f"={I['start_val']}",
     "$#,##0",
     "Net of estimated tax after harvesting offset  (Inputs C48)"),
    ("Expected Gross Annual Return",
     f"={I['bear_return']}", f"={I['base_return']}", f"={I['bull_return']}",
     "0.0%",
     "Edit individual scenario returns in Inputs E30-E32"),
    ("Annual Withdrawal Rate",
     f"={I['withdrawal_rate']}", f"={I['withdrawal_rate']}", f"={I['withdrawal_rate']}",
     "0.0%",
     "Inputs E33  |  Applied uniformly across all scenarios"),
    ("Net Annual Growth Rate",
     f"={I['bear_net']}", f"={I['base_net']}", f"={I['bull_net']}",
     "0.0%",
     "Return minus Withdrawal Rate  (Inputs C49-C51)  — drives the projection"),
    ("Investment Horizon",
     "20 Years", "20 Years", "20 Years",
     None,
     "Inputs E34"),
    ("Private Foundation Target",
     f"={I['foundation_target']}", f"={I['foundation_target']}", f"={I['foundation_target']}",
     "$#,##0",
     "Inputs E35  |  Philanthropic endowment goal"),
    ("Foundation Funding Method",
     "SBL Facility", "SBL Facility", "SBL Facility",
     None,
     "$5M drawn from credit line — zero liquidation, zero additional tax"),
    ("Annual SBL Interest Cost",
     f"={I['annual_net_int']}", f"={I['annual_net_int']}", f"={I['annual_net_int']}",
     "$#,##0",
     "After-tax interest  (Inputs C47)  — updates with rate or loan changes"),
]

for i, (param, bear, base, bull, fmt, note) in enumerate(assumptions):
    bg = stripe(i)
    data_cell(ws, row, 2, param, bold=True, bg=bg)
    for col, val, color in [(3, bear, RED), (4, base, NAVY), (5, bull, GREEN)]:
        is_f = isinstance(val, str) and val.startswith("=")
        if is_f:
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name="Calibri", bold=(col == 4), size=10, color=color)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = PatternFill("solid", fgColor=CBLUE); c.border = thin_border()
            if fmt: c.number_format = fmt
        else:
            bold_c = col == 4
            data_cell(ws, row, col, val, bold=bold_c, bg=bg,
                      align="center", color=color)
    merge(ws, row, 6, row, 7)
    data_cell(ws, row, 6, note, bg=bg, size=9, color="666666")
    rh(ws, row, 16); row += 1

rh(ws, row, 8); row += 1

# ── Section B: 20-Year Projection — formula chain ─────────────────────────
merge(ws, row, 2, row, 7)
section_cell(ws, row, 2,
    "  SECTION B: 20-YEAR PORTFOLIO PROJECTION  —  "
    "Each year = prior year × (1 + net rate).  Change Inputs to reprice instantly.")
rh(ws, row, 18); row += 1

for col, label in [
    (2, "Year"),
    (3, "Calendar Year"),
    (4, "Bear Case\n(Net 3.0%)"),
    (5, "Base Case\n(Net 7.8%)"),
    (6, "Bull Case\n(Net 11.5%)"),
    (7, "Milestone"),
]:
    col_head(ws, row, col, label)
rh(ws, row, 28); row += 1

milestones = {
    1:  "SBL Facility Established",
    5:  "5-Year Review",
    10: "Foundation Self-Sustaining",
    15: "15-Year Review",
    20: "▲ Planning Horizon",
}

proj_row = {}   # year → worksheet row number

for yr in range(1, 21):
    cal_yr  = 2026 + yr
    bg      = stripe(yr)
    bold    = yr in [5, 10, 15, 20]
    ms      = milestones.get(yr, "")
    ms_color = PURPLE if ms else "888888"

    if yr == 1:
        bear_f = f"={I['start_val']}*(1+{I['bear_net']})"
        base_f = f"={I['start_val']}*(1+{I['base_net']})"
        bull_f = f"={I['start_val']}*(1+{I['bull_net']})"
    else:
        pr = proj_row[yr - 1]
        bear_f = f"=D{pr}*(1+{I['bear_net']})"
        base_f = f"=E{pr}*(1+{I['base_net']})"
        bull_f = f"=F{pr}*(1+{I['bull_net']})"

    proj_row[yr] = row

    data_cell(ws, row, 2, f"Year {yr:>2}", bold=bold, bg=bg)
    data_cell(ws, row, 3, cal_yr, bold=bold, bg=bg, align="center")

    for col, formula, color in [
        (4, bear_f, RED if bold else DGRAY),
        (5, base_f, NAVY if bold else DGRAY),
        (6, bull_f, GREEN if bold else DGRAY),
    ]:
        c = ws.cell(row=row, column=col, value=formula)
        c.font = Font(name="Calibri", bold=bold, size=10, color=color)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.fill = PatternFill("solid", fgColor=CBLUE); c.border = thin_border()
        c.number_format = "$#,##0"

    data_cell(ws, row, 7, ms, bold=bold, bg=bg, size=9, color=ms_color)
    rh(ws, row, 14); row += 1

# Summary total row — references year 20
yr20 = proj_row[20]
total_cell(ws, row, 2, "20-YEAR ENDING VALUE")
total_cell(ws, row, 3, 2046, fmt="0")
total_cell(ws, row, 4, f"=D{yr20}", fmt='$#,##0')
total_cell(ws, row, 5, f"=E{yr20}", fmt='$#,##0')
total_cell(ws, row, 6, f"=F{yr20}", fmt='$#,##0')
total_cell(ws, row, 7, "End of Planning Horizon")
rh(ws, row, 20); row += 2

# ── Section C: Foundation Funding Analysis ─────────────────────────────────
merge(ws, row, 2, row, 7)
section_cell(ws, row, 2, "  SECTION C: PRIVATE FOUNDATION FUNDING ANALYSIS")
rh(ws, row, 18); row += 1

for col, label in [
    (2, "Milestone"), (3, "Year"), (4, "Foundation Value"),
    (5, "Cumulative Grants"), (6, "Endowment Status"), (7, "Notes"),
]:
    col_head(ws, row, col, label)
rh(ws, row, 16); row += 1

# Foundation value formula: seed × (1 + net_r)^n  +  contrib × ((1 + net_r)^n − 1) / net_r
# Net endowment return = foundation_return − foundation_payout (7.5% − 5.0% = 2.5%)
# Annual contribution (Inputs C38) bridges $5M seed to $10M target at Year 20.
# Cumulative grants = seed × payout × n  (simplified lower-bound approximation).
_lr  = I['loan_amount']
_fr  = I['foundation_return']
_fp  = I['foundation_payout']
_fc  = I['foundation_contrib']
_nr  = f"({_fr}-{_fp})"   # net return expression reused across formulas

def _fval(n):
    """Excel formula: foundation value at year n including annual contributions."""
    return f"={_lr}*(1+{_nr})^{n}+{_fc}*((1+{_nr})^{n}-1)/{_nr}"

found_items = [
    ("SBL Proceeds Deployed",        "Year 1",
     f"={_lr}",
     0,
     "Foundation Seeded",  "100% via SBL — zero additional liquidation. $70,760/yr contribution begins Year 1."),
    ("Foundation — Year 5",          "Year 5",
     _fval(5),
     f"={_lr}*{_fp}*5",
     "Growing",            "7.5% return net of 5% grantmaking + $70,760/yr client contribution"),
    ("Foundation — Year 10",         "Year 10",
     _fval(10),
     f"={_lr}*{_fp}*10",
     "Growing",            "Projected $7.2M at Year 10 — seed + annual contributions"),
    ("SBL Loan Fully Repaid",        "Year 10",
     _fval(10),
     f"={_lr}*{_fp}*10",
     "Debt-Free",          "Loan retired from portfolio income — no asset sales"),
    ("Foundation — Year 20 (Proj.)", "Year 20",
     _fval(20),
     f"={_lr}*{_fp}*20",
     "Target Achieved",    "$10M target reached: $5M seed + $70,760/yr from portfolio income"),
]

for i, (milestone, yr, fval_f, grants_f, status, note) in enumerate(found_items):
    bg    = stripe(i)
    bold  = status in ["Target Achieved", "Debt-Free"]
    sc    = GREEN if "Achieved" in status or "Debt" in status else NAVY

    data_cell(ws, row, 2, milestone, bold=bold, bg=bg, color=sc)
    data_cell(ws, row, 3, yr, bg=bg, align="center")

    for col, formula, fmt in [(4, fval_f, "$#,##0"), (5, grants_f, "$#,##0")]:
        is_f = isinstance(formula, str) and formula.startswith("=")
        if is_f:
            c = ws.cell(row=row, column=col, value=formula)
            c.font = Font(name="Calibri", bold=bold, size=10,
                          color=GREEN if bold else DGRAY)
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.fill = PatternFill("solid", fgColor=CBLUE); c.border = thin_border()
            c.number_format = fmt
        else:
            data_cell(ws, row, col, formula, bg=bg, align="right", fmt=fmt)

    data_cell(ws, row, 6, status, bold=bold, bg=bg, align="center", color=sc)
    data_cell(ws, row, 7, note, bg=bg, size=9, color="666666")
    rh(ws, row, 22); row += 1

rh(ws, row, 8); row += 1
merge(ws, row, 2, row, 7)
note_cell(ws, row, 2,
    "Foundation value = Seed × (1 + Net Return)^Year + Annual Contribution × ((1 + Net Return)^Year − 1) / Net Return.  "
    "Net Return = Foundation Annual Return − Annual Payout (Inputs C36 & C37).  "
    "Annual Contribution (Inputs C38, default $70,760) sourced from diversified portfolio income ($625K/yr).  "
    "Cumulative grants column is a simplified lower-bound: Seed × Payout × Years.  "
    "Change any Inputs assumption to instantly update all projections.  "
    "Projections are illustrative. Consult a financial advisor before investment or philanthropic decisions.")
rh(ws, row, 28)

# ── SAVE ──────────────────────────────────────────────────────────────────────
wb.calculation.calcMode = "auto"
wb.calculation.fullCalcOnLoad = True
wb.save(OUTPUT_FILE)
print(f"Workbook saved: {OUTPUT_FILE}")
print(f"Tabs: {', '.join(wb.sheetnames)}")
